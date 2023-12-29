import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

INDENT = '\t'

# 再帰的にコンテンツを解析し、Excelシートに書き込む関数
def parse_content(sheet, depth, content):
    name = content.get('name', '')

    # フォルダの場合は名前の末尾に「/」を追加し、フォントを設定する
    if content.get('type') == 'directory':
        name += '/'  # フォルダ名の場合は末尾に「/」を追加
        font = Font(bold=True, size=14)  # フォルダ名のフォントサイズを14、かつ太字に設定
    else:
        font = Font(size=11)  # ファイル名のフォントサイズを11に設定

    # インデントと名前をシートに追加する行の形式に変換
    row = [''] * depth + [name]
    sheet.append(row)  # シートに行を追加

    # 追加したセルにフォントを適用する
    cell = sheet.cell(row=sheet.max_row, column=len(row))
    cell.font = font

    # フォルダの場合は再帰的に内容を解析
    if content.get('type') == 'directory':
        for c in content.get('contents', []):
            parse_content(sheet, depth + 1, c)

# JSONをExcelに変換する関数
def json_to_excel():
    try:
        contents = json.load(sys.stdin)  # 標準入力からJSONデータを読み込む
    except Exception as e:
        print(e)  # 例外が発生した場合はエラーメッセージを表示
        return  

    # 新しいExcelブックを作成し、アクティブなシートを設定
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'シート１'  # シート名を設定

    # コンテンツを解析してExcelに書き込む
    for content in contents:
        parse_content(sheet, 0, content)

    wb.save('output.xlsx')  # Excelファイルを保存
    
    print("Excel file 'output.xlsx' created successfully.")

# メイン処理
if __name__ == '__main__':
    json_to_excel()
